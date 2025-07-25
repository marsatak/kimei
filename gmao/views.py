import io
import json
import logging
import os
import re
from datetime import datetime

from aiohttp.web_fileresponse import FileResponse
from aiohttp.web_response import json_response
from django.contrib import messages
from django.core import serializers
from django.urls import reverse
from django.core.cache import cache
from django.utils.archive import extract
from openpyxl.reader.excel import load_workbook
from rest_framework.decorators import api_view
from rest_framework import status
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse, Http404
from rest_framework.response import Response
from rest_framework.parsers import JSONParser
from django.shortcuts import render, redirect
from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction
from django.views.decorators.cache import cache_control
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q, F
from django.db.models.functions import ExtractYear
from django.core.paginator import Paginator
from rest_framework.decorators import api_view
from django.db.models.functions import Substr, StrIndex
from django.db.models import Prefetch
from django.utils import timezone
from datetime import timedelta
from datetime import datetime, time
from django.db.models import CharField, Value as V
from django.db.models.functions import Substr
from openpyxl import Workbook

from django.utils.dateparse import parse_datetime

from accounts.models import Employee, AbstractUser
from gmao.forms import DoleanceForm
from gmao.models import (
    Appelant, Client, Station, Doleance, Intervention, InterventionPersonnel,
    Poste, Personnel, Pointage,
    Boutique, Compresseur, Cuve, Piste, AppareilDistribution, Servicing, Elec, GroupeElectrogene,
    Auvent, Totem, Produit, Pistolet)
from gmao.serializers import (
    ClientSerializer, StationSerializer, DoleanceSerializer, InterventionSerializer,
    PosteSerializer, PersonnelSerializer, PointageSerializer,
    AppareilDistributionSerializer
)
from django.views.decorators.http import require_http_methods

from kimei import settings
from .models import AppareilDistribution, Station, ModeleAd
from django.db.models import Prefetch
from .models import Pistolet
from django.db import IntegrityError
from dateutil.relativedelta import relativedelta
from gmao_teams.models import EquipePersonnel, DoleanceEquipe, Equipe
from .models import Piece, UnitePiece, TypePiece
from gmao_teams.models import Equipe
from .utils import filter_active_doleances
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponseBadRequest
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponseBadRequest
from django.views.decorators.http import require_http_methods
from django.db.models import Min, Max
from django.core.exceptions import ValidationError
from django.db import DatabaseError
from django.db.models.functions import ExtractYear
from django.db.models.functions import TruncYear
from django.http import JsonResponse
from django.core.serializers import serialize
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models import Prefetch
from django.views.decorators.cache import cache_control
from django.db import models
from django.db.models import Min, Max
from dateutil.relativedelta import relativedelta
import logging

# Ajoutez ces imports pour WhatsApp et SMS
# from twilio.rest import Client

logger = logging.getLogger(__name__)


# #####################Test######################
@require_http_methods(["GET", "POST"])
def test_view(request):
    print('post')
    return JsonResponse({'message': 'Test réussi'})


# #####################TimeStamp######################
def api_data(request):
    last_update = request.GET.get('last_update')

    if last_update:
        last_update = timezone.datetime.fromtimestamp(float(last_update))
        doleances = Doleance.objects.filter(modified_at__gt=last_update)
        interventions = Intervention.objects.filter(modified_at__gt=last_update)
        personnel = Personnel.objects.filter(modified_at__gt=last_update)
    else:
        doleances = Doleance.objects.all()
        interventions = Intervention.objects.all()
        personnel = Personnel.objects.all()

    has_new_data = doleances.exists() or interventions.exists() or personnel.exists()

    data = {
        'hasNewData': has_new_data,
        'timestamp': timezone.now().timestamp(),
        'data': {
            'doleances': list(doleances.values()),
            'interventions': list(interventions.values()),
            'personnel': list(personnel.values()),
        }
    }

    return JsonResponse(data)


# ##################### Début Test connexion bdd pas obligatoire ######################
def test_db_connection(request):
    try:
        count = Doleance.objects.using('kimei_db').count()
        return HttpResponse(f"Nombre de doléances : {count}")
    except Exception as e:
        return HttpResponse(f"Erreur : {str(e)}")


# ##################### Fin Test connexion bdd pas obligatoire ######################

# ##################### Début Dashboard ######################
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def home(request):
    form = DoleanceForm()
    context = {
        'form': form,
        'user_role': request.user.role,
    }
    # print(request.user.role)

    # if request.user.role == 'ADMIN':
    #     doleances = Doleance.objects.using('kimei_db').all()
    #     techniciens = Employee.objects.filter(role='TECH', statut='PRS')
    #     context.update({
    #         'doleances': doleances,
    #         'techniciens': techniciens,
    #     })
    #     equipes = Equipe.objects.using('teams_db').all()
    #     equipes_data = []
    #
    #     for equipe in equipes:
    #         doleance_ids = DoleanceEquipe.objects.using('teams_db').filter(equipe=equipe).values_list('doleance_id',
    #                                                                                                   flat=True)
    #         doleances = Doleance.objects.using('kimei_db').filter(id__in=doleance_ids).exclude(statut='TER')
    #
    #         equipes_data.append({
    #             'equipe': equipe,
    #             'doleances': doleances
    #         })
    #
    #     context['equipes_data'] = equipes_data
    if request.user.role == 'ADMIN':
        equipes = Equipe.objects.using('teams_db').all()
        equipes_data = []

        for equipe in equipes:
            # Récupérer les IDs des doléances pour cette équipe
            doleance_ids = list(DoleanceEquipe.objects.using('teams_db')
                                .filter(equipe=equipe)
                                .values_list('doleance_id', flat=True))

            # Récupérer les doléances correspondantes
            doleances = Doleance.objects.using('kimei_db').filter(id__in=doleance_ids)
            doleances = filter_active_doleances(doleances)

            # Récupérer les IDs des techniciens pour cette équipe
            technicien_ids = list(EquipePersonnel.objects.using('teams_db')
                                  .filter(equipe=equipe)
                                  .values_list('personnel_id', flat=True))

            # Récupérer les techniciens correspondants
            techniciens = Personnel.objects.using('kimei_db').filter(id__in=technicien_ids)

            equipes_data.append({
                'id': equipe.id,
                'nom': equipe.nom,
                'description': equipe.description,
                'doleances': [
                    {
                        'id': d.id,
                        'ndi': d.ndi,
                        'panne_declarer': d.panne_declarer,
                        'statut': d.statut,
                        'date_transmission': d.date_transmission,
                        'station': d.station.libelle_station if d.station else None
                    } for d in doleances
                ],
                'techniciens': [
                    {
                        'id': t.id,
                        'nom': t.nom_personnel,
                        'prenom': t.prenom_personnel,
                        'statut': t.statut
                    } for t in techniciens
                ]
            })

        context['equipes_data'] = equipes_data
    elif request.user.role == 'TECH':
        personnel = Personnel.objects.using('kimei_db').get(matricule=request.user.matricule)
        equipe = EquipePersonnel.objects.using('teams_db').filter(personnel_id=personnel.id).first()
        if equipe:
            doleance_ids = DoleanceEquipe.objects.using('teams_db').filter(equipe=equipe.equipe).values_list(
                'doleance_id', flat=True)
            doleances = Doleance.objects.using('kimei_db').filter(id__in=doleance_ids)
            context.update({
                'doleances': doleances,
            })

    return render(request, 'gmao/home.html', context)


# ##################### Début Liste des dléances en cours ######################

@api_view(['GET'])
def getDoleanceEncours(request):
    try:
        doleances = (
            (Doleance.objects.all()).using('kimei_db')
            .exclude(statut='TER')
            .order_by('-date_transmission')
        )
        print(datetime.now().month)
        print(datetime.now().year)
        print(doleances)
        doleances_data = []
        for doleance in doleances:
            doleance_dict = DoleanceSerializer(doleance).data
            buttons = ''
            if doleance.statut == 'NEW':
                buttons += f'<button class="btn btn-primary btn-sm prendre-en-charge" data-id="{doleance.id}">Prendre en charge</button> '
            elif doleance.statut == 'ATT':
                intervention = Intervention.objects.filter(doleance=doleance, is_done=False).first()
                if intervention:
                    buttons += f'<button class="btn btn-success btn-sm commencer-intervention" data-id="{intervention.id}">Commencer</button> '
            elif doleance.statut == 'INT':
                intervention = Intervention.objects.filter(doleance=doleance, is_done=False).first()
                if intervention:
                    buttons += f'<button class="btn btn-warning btn-sm terminer-intervention" data-id="{intervention.id}">Terminer</button> '
            doleance_dict['actions'] = buttons
            doleances_data.append(doleance_dict)

        return Response(doleances_data)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ##################### Fin Liste des dléances en cours ######################


# ##################### Début Equipes et leurs doléances ######################


@api_view(['GET'])
def get_equipes_data(request):
    equipes = Equipe.objects.all()
    data = []
    for equipe in equipes:
        # Récupérer les IDs des personnels et des doléances associés à cette équipe
        personnel_ids = equipe.get_personnel_ids()
        doleance_ids = equipe.get_doleance_ids()

        # Récupérer les techniciens et les doléances correspondants
        techniciens = Personnel.objects.using('kimei_db').filter(id__in=personnel_ids)
        doleances = Doleance.objects.using('kimei_db').filter(id__in=doleance_ids)

        equipe_data = {
            'id': equipe.id,
            'nom': equipe.nom,
            'description': equipe.description,
            'techniciens': [
                {
                    'nom': tech.nom_personnel,
                    'prenom': tech.prenom_personnel,
                    'statut': tech.statut
                } for tech in techniciens
            ],
            'doleances': [
                {
                    'ndi': dol.ndi,
                    'station': dol.station.libelle_station if dol.station else None,
                    'panne_declarer': dol.panne_declarer,
                    'statut': dol.statut
                } for dol in doleances
            ]
        }
        data.append(equipe_data)
    return JsonResponse(data, safe=False)


# ##################### Fin Equipes et leurs doléances ######################
@api_view(['GET'])
def getPoste(request):
    postes = Poste.objects.all()
    postes_serializer = PosteSerializer(postes, many=True)
    return Response(postes_serializer.data, content_type='application/json; charset=UTF-8')


# #####################Pointage Arrivée######################
@require_POST
def mark_arrivee(request, personnel_id):
    try:
        personnel = Personnel.objects.get(id=personnel_id)
        personnel.statut = 'PRS'
        personnel.save()
        return JsonResponse({'success': True, 'message': 'Arrivée marquée avec succès'})
    except Personnel.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Personnel non trouvé'}, status=404)


# #####################Pointage Départ######################

@require_POST
def mark_depart(request, personnel_id):
    try:
        personnel = Personnel.objects.get(id=personnel_id)
        personnel.statut = 'ABS'
        personnel.save()
        return JsonResponse({'success': True, 'message': 'Départ marqué avec succès'})
    except Personnel.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Personnel non trouvé'}, status=404)


# #####################Début Liste Personnel######################
@api_view(['GET'])
def getPersonnel(request):
    try:
        personnels = Personnel.objects.filter(is_active=True).select_related('poste').order_by('-poste_id')
        data = [{
            'id': p.id,
            'nom_personnel': p.nom_personnel,
            'prenom_personnel': p.prenom_personnel,
            'statut': p.statut,
            'poste': {
                'id': p.poste.id,
                'nom_poste': p.poste.nom_poste
            }
        } for p in personnels]
        return Response(data, content_type='application/json; charset=UTF-8')
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# #####################Fin Liste Personnel######################

# #####################Début Maj Statut Personnel######################
@api_view(['PUT'])
def updatePersonnel(request, id):
    personnels = Personnel.objects.get(pk=id)
    personnels_serializer = PersonnelSerializer(personnels, data=request.data)
    if personnels_serializer.is_valid():
        personnels_serializer.save()
        return Response(personnels_serializer.data, content_type='application/json; charset=UTF-8')
    else:
        return Response(personnels_serializer.errors)


# #####################Fin Maj Statut Personnel######################

# #####################Liste Pointage######################
@api_view(['GET'])
def getPointage(request):
    pointages = (Pointage.objects.all()
                 .filter(date_heure_arrive__day=datetime.now().day)
                 .filter(date_heure_arrive__month=datetime.now().month)
                 .filter(date_heure_arrive__year=datetime.now().year))

    pointages_serializer = PointageSerializer(pointages, many=True)
    return Response(pointages_serializer.data)


# #####################Faire Un Pointage######################
@api_view(['POST'])
def postPointage(request):
    data = request.data
    personnel_id = data['personnel_id']
    today = timezone.now().date()

    existing_pointage = Pointage.objects.filter(
        personnel_id=personnel_id,
        date_heure_arrive__date=today
    ).first()

    if existing_pointage:
        return Response({'message': 'Un pointage existe déjà pour aujourd\'hui'}, status=status.HTTP_400_BAD_REQUEST)

    pointage = Pointage.objects.create(
        personnel_id=personnel_id,
        date_heure_arrive=timezone.now()
    )
    serializer = PointageSerializer(pointage)
    return Response(serializer.data, status=status.HTTP_201_CREATED)


@api_view(['PUT'])
def putPointage(request, pk):
    serializer = PointageSerializer(data=request.data)
    try:
        personnel = (Personnel.objects
                     .filter(pointage__date_heure_arrive__day=datetime.now().day)
                     .get(id=pk))
    except Personnel.DoesNotExist:
        return Response({'error': 'Personnel non trouvé'}, status=status.HTTP_404_NOT_FOUND)
    print(pk)
    print(personnel)
    pointages = Pointage.objects.filter(personnel=personnel)
    pointage = pointages.get(date_heure_arrive__day=datetime.now().day)
    print(pointage)
    if pointage.date_heure_sortie is None:
        pointage.date_heure_sortie = datetime.now()
        diff = ((pointage.date_heure_sortie - pointage.date_heure_arrive)
                .total_seconds())
        pointage.seconde_actif = int(diff)
        serializer = PointageSerializer(pointage, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# #####################Début Réinitialiser Pointage######################
def resetPointage(request):
    personnels = Personnel.objects.all().filter(is_active=True)
    for personnel in personnels:
        personnel.statut = "ABS"
        personnel.save()
    messages.info(request, 'Déconnexion réussie')
    return redirect('accounts:logout-admin')


# #####################Fin Réinitialiser Pointage######################

# #####################Début Création doléance######################
@csrf_exempt
@require_http_methods(["POST"])
def create_doleance(request):
    if request.method == 'POST':
        form = DoleanceForm(request.POST)
        if form.is_valid():
            doleance = form.save(commit=False)
            doleance.date_transmission = timezone.now()

            # doleance.date_deadline = form.cleaned_data['date_deadline']
            # date_deadline = form.cleaned_data.get('date_deadline')
            date_deadline = form.cleaned_data.get('date_deadline')
            if not date_deadline:
                # Convertir la chaîne de date en objet datetime
                date_deadline = timezone.now() + timezone.timedelta(days=1)

            doleance.date_deadline = date_deadline
            doleance.element = form.cleaned_data['element']
            doleance.panne_declarer = form.cleaned_data['panne_declarer'].upper()
            doleance.statut = "NEW"
            doleance.save()

            return JsonResponse({
                'success': True,
                'message': 'Doléance créée avec succès',
                'id': doleance.id
            })
        else:
            return JsonResponse({
                'success': False,
                'errors': form.errors
            })
    else:
        return JsonResponse({
            'success': False,
            'message': 'Méthode non autorisée'
        }, status=405)


# #####################Fin Création doléance######################

# #####################Début Liste des doléances######################
@require_GET
def get_doleance(request, doleance_id):
    try:
        doleance = get_object_or_404(Doleance, id=doleance_id)
        data = {
            'id': doleance.id,
            'client': doleance.station.client.id,
            'station': doleance.station.id,
            'appelant': doleance.appelant.id if doleance.appelant else None,
            'date_transmission': doleance.date_transmission,
            'type_transmission': doleance.type_transmission,
            'panne_declarer': doleance.panne_declarer,
            'element': doleance.element,
            'commentaires': doleance.commentaire,
            'type_contrat': doleance.type_contrat,
            'date_deadline': timezone.localtime(doleance.date_deadline).strftime(
                '%d/%m/%Y %H:%M') if doleance.date_deadline else '',
            # 'date_deadline': doleance.date_deadline.strftime('%d/%m/%Y %H:%M') if doleance.date_deadline else '',
        }
        return JsonResponse({'success': True, 'doleance': data})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# #####################Fin Liste des doléances#####################

# #####################Début Procédure de maj des doléances######################
@csrf_exempt
@require_POST
def update_doleance(request, doleance_id):
    try:
        doleance = get_object_or_404(Doleance, id=doleance_id)
        # logger.info(f"Tentative de mise à jour de la doléance {doleance_id}")
        # logger.info(f"Données reçues : {request.POST}")
        current_statut = doleance.statut
        contrat = doleance.type_contrat

        form = DoleanceForm(request.POST, instance=doleance)
        if form.is_valid():
            # logger.info("Formulaire valide, sauvegarde en cours")
            updated_doleance = form.save(commit=False)

            updated_doleance.statut = current_statut

            # Accédez aux données nettoyées seulement si le formulaire est valide
            updated_doleance.panne_declarer = form.cleaned_data['panne_declarer'].upper()
            updated_doleance.commentaire = form.cleaned_data.get('commentaire', '')

            # Gestion de la transmission
            if 'date_transmission' in form.cleaned_data and form.cleaned_data['date_transmission']:
                date_transmission = form.cleaned_data['date_transmission']
                if timezone.is_naive(date_transmission):
                    updated_doleance.date_transmission = timezone.make_aware(date_transmission,
                                                                             timezone.get_current_timezone())
                else:
                    updated_doleance.date_transmission = timezone.localtime(date_transmission)

            # Regeneration NDI
            if contrat != updated_doleance.type_contrat:
                updated_doleance.ndi = (doleance.ndi[:-1]) + updated_doleance.type_contrat

            # Gestion de la date_deadline
            if 'date_deadline' in form.cleaned_data and form.cleaned_data['date_deadline']:
                date_deadline = form.cleaned_data['date_deadline']
                if timezone.is_naive(date_deadline):
                    updated_doleance.date_deadline = timezone.make_aware(date_deadline, timezone.get_current_timezone())
                else:
                    updated_doleance.date_deadline = timezone.localtime(date_deadline)

            updated_doleance.save()
            # logger.info(f"Doléance {doleance_id} mise à jour avec succès")
            return JsonResponse({
                'success': True,
                'message': 'Doléance mise à jour avec succès',
                'id': updated_doleance.id,
                'statut': updated_doleance.statut,
                'element': updated_doleance.element
            })
        else:
            # logger.error(f"Erreurs de formulaire : {form.errors}")
            return JsonResponse({
                'success': False,
                'errors': form.errors
            })
    except Exception as e:
        # logger.exception(f"Erreur lors de la mise à jour de la doléance {doleance_id}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


# #####################Fin Procédure de maj des doléances######################
# #####################Début Liste Stations######################

def load_stations(request):
    client_id = request.GET.get('client')
    if not client_id:
        return JsonResponse([], safe=False)
    try:
        stations = Station.objects.filter(client_id=client_id).order_by('libelle_station')
        return JsonResponse(list(stations.values('id', 'libelle_station')), safe=False)
    except ValueError:
        return JsonResponse({'error': 'ID client invalide'}, status=400)
    except Exception as e:
        logger.error(f"Erreur lors du chargement des stations: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


# #####################Fin Liste Stations######################
# ##################### Début chargement liste stations ######################
def search_stations(request):
    client_id = request.GET.get('client_id')
    search_term = request.GET.get('term', '')

    stations = Station.objects.filter(client_id=client_id)
    if search_term:
        stations = stations.filter(Q(libelle_station__icontains=search_term) | Q(lieu_station__icontains=search_term))

    data = list(stations.values('id', 'libelle_station', 'lieu_station'))
    print("Station data:", data)
    return JsonResponse(data, safe=False)


# ##################### Fin chargement liste stations ######################

# ##################### Début Liste des appelants ######################
def load_appelants(request):
    client_id = request.GET.get('client')
    if not client_id:
        return JsonResponse([], safe=False)
    logger.info(f"Tentative de chargement des appelants pour le client_id: {client_id}")
    try:
        client_id = int(client_id)
        appelants = Appelant.objects.filter(client_id=client_id).order_by('nom_appelant')
        logger.info(f"Nombre d'appelants trouvés: {appelants.count()}")
        data = list(appelants.values('id', 'nom_appelant', 'prenom_appelant'))
        return JsonResponse(data, safe=False)
    except DatabaseError as e:
        logger.error(f"Erreur de base de données lors du chargement des appelants: {str(e)}")
        return JsonResponse({'error': 'Erreur de base de données'}, status=500)
    except Exception as e:
        logger.error(f"Erreur inattendue lors du chargement des appelants: {str(e)}")
        return JsonResponse({'error': 'Erreur serveur inattendue'}, status=500)


# ##################### Fin Liste des appelants ######################

# ##################### Début Liste des éléments ######################
def load_elements(request):
    station_id = request.GET.get('station')
    if not station_id:
        return JsonResponse([], safe=False)

    print(f"Station ID: {station_id}")
    elements = DoleanceForm.get_station_elements(station_id)
    return JsonResponse(elements)


# ##################### Fin Liste des éléments ######################
# ##################### Fin Dashboard ######################

# ##################### Début Commencer une intervention ######################
@require_http_methods(["GET", "POST"])
def commencer_intervention(request, intervention_id):
    try:
        intervention = get_object_or_404(Intervention, id=intervention_id)
        kilometrage = request.POST.get('kilometrage')
        if intervention.etat_doleance != 'ATT':
            return JsonResponse({'success': False, 'message': 'L\'intervention n\'est pas en attente'})

        if not intervention.top_debut:
            intervention.top_debut = timezone.now()
            intervention.is_half_done = True
            intervention.etat_doleance = 'INT'
            intervention.kilometrage_depart_debut = kilometrage
            intervention.save()

        doleance = intervention.doleance
        doleance.statut = 'INT'
        doleance.save()

        intervention_personnels = InterventionPersonnel.objects.filter(intervention=intervention)
        for intervention_personnel in intervention_personnels:
            technicien = intervention_personnel.personnel
            technicien.statut = 'INT'
            technicien.save()

        return JsonResponse({
            'success': True,
            'message': 'Intervention déclenchée avec succès',
            'intervention_id': intervention.id,
            'top_debut': intervention.top_debut.strftime('%Y-%m-%d %H:%M:%S')
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


# ##################### Fin Commencer une intervention ######################

# ##################### Début Liste des techs disponibles ######################
@require_GET
def get_techniciens_disponibles(request):
    techniciens = Personnel.objects.filter(statut='PRS')
    return JsonResponse({
        'success': True,
        'techniciens': list(techniciens.values('id', 'nom_personnel', 'prenom_personnel'))
    })


# ##################### Fin Liste des techs disponibles ######################

# #####################Déclencher une intervention######################
@require_POST
def declencher_intervention(request, doleance_id):
    doleance = get_object_or_404(Doleance, id=doleance_id)

    if doleance.statut not in ['NEW', 'ATP', 'ATD']:
        return JsonResponse({'success': False, 'message': 'Cette doléance ne peut pas être traitée'})

    data = json.loads(request.body)
    techniciens_ids = data.get('techniciens', [])

    intervention = Intervention.objects.create(
        doleance=doleance,
        top_depart=timezone.now(),
        is_done=False,
        is_half_done=False,
        is_going_home=False,
        etat_doleance='ATT'
    )

    doleance.statut = 'ATT'
    doleance.date_debut = timezone.now()
    doleance.save()

    for tech_id in techniciens_ids:
        technicien = Personnel.objects.get(id=tech_id)
        InterventionPersonnel.objects.create(intervention=intervention, personnel=technicien)
        technicien.statut = 'ATT'  # Le technicien est en attente
        technicien.save()

    return JsonResponse({
        'success': True,
        'message': 'Intervention déclenchée avec succès',
        'intervention_id': intervention.id
    })


# #####################Annuler une intervention######################
@require_POST
def annuler_intervention(request, intervention_id):
    try:
        intervention = get_object_or_404(Intervention, id=intervention_id)
        doleance = intervention.doleance
        ancien_statut = doleance.statut

        intervention_personnels = InterventionPersonnel.objects.filter(intervention=intervention)
        for intervention_personnel in intervention_personnels:
            technicien = intervention_personnel.personnel
            technicien.statut = 'PRS'
            technicien.save()
        if ancien_statut in ['ATP', 'ATD']:
            doleance.statut = ancien_statut
        else:
            doleance.statut = 'NEW'
        doleance.save()
        # Réinitialiser l'intervention
        intervention.delete()

        return JsonResponse({
            'success': True,
            'message': 'Intervention annulée avec succès',
            'nouveau_statut': doleance.statut
        })
    except Intervention.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Intervention non trouvée'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


# #####################Début Liste des interventions######################
@login_required
def liste_interventions(request):
    current_date = timezone.now().date()
    context = {
        'current_date': current_date,
        'user_role': request.user.role,
        'is_admin': request.user.role == 'ADMIN'
    }
    return render(request, 'gmao/liste_interventions.html', context)


# #####################Début Détails des interventions######################
@login_required
def detail_intervention(request, intervention_id):
    intervention = get_object_or_404(Intervention, id=intervention_id)
    techniciens = InterventionPersonnel.objects.filter(intervention=intervention).select_related('personnel')

    # Vérifier si l'intervention est pour aujourd'hui
    is_today = intervention.top_depart.date() == timezone.now().date()

    # Vérifier si le technicien connecté est assigné à cette intervention
    user_assigned = techniciens.filter(
        personnel__matricule=request.user.matricule).exists() if request.user.role == 'TECH' else True

    context = {
        'intervention': intervention,
        'techniciens': techniciens,
        'is_today': is_today,
        'user_assigned': user_assigned,
        'heure_actuelle': timezone.now().strftime('%H:%M'),
    }
    return render(request, 'gmao/detail_intervention.html', context)


# #####################Début Procédure de clôture######################
def capitalizeSentences(string):
    if not string:
        return string
    # Divise la chaîne en phrases
    sentences = re.split(r'([.!?]\s*)', string)

    # Capitalise la première lettre de chaque phrase
    capitalized = []
    for i in range(0, len(sentences), 2):
        sentence = sentences[i].strip()
        if sentence:
            sentence = sentence[0].upper() + sentence[1:]
        if i + 1 < len(sentences):
            sentence += sentences[i + 1]  # Ajoute la ponctuation
        capitalized.append(sentence)
    return ''.join(capitalized)


@csrf_exempt
@require_POST
def terminer_travail(request, intervention_id):
    global date
    try:
        intervention = get_object_or_404(Intervention, id=intervention_id)
        if not intervention.numero_fiche:
            isModification = False
        else:
            isModification = True
        intervention.description_panne = capitalizeSentences(request.POST.get('description_panne', ''))
        intervention.resolution = capitalizeSentences(request.POST.get('resolution', ''))
        intervention.observations = capitalizeSentences(request.POST.get('observations', ''))
        intervention.pieces_changees = capitalizeSentences(request.POST.get('pieces_changees', ''))
        # new_element = capitalizeSentences(request.POST.get('element', ''))

        statut_final = request.POST.get('statut_final', '').upper()

        if not statut_final or statut_final not in ['TER', 'ATP', 'ATD']:
            return JsonResponse({'success': False, 'message': f'Statut final invalide: {statut_final}'})

        # Traitement du numéro de fiche
        numero_fiche = request.POST.get('numero_fiche', '')
        try:
            numero_fiche = int(numero_fiche)
            if numero_fiche < 0 or numero_fiche > 99999:
                return JsonResponse({'success': False, 'message': 'Le numéro de fiche doit être entre 0 et 99999'})

            numero_fiche_complet = f"00{numero_fiche:05d}"

            # Vérifier si le numéro de fiche existe déjà
            if Intervention.objects.filter(numero_fiche=numero_fiche_complet).exclude(id=intervention_id).exists():
                return JsonResponse({
                    'success': False,
                    'message': 'Ce numéro de fiche est déjà utilisé. Veuillez en choisir un autre.',
                    'error_type': 'numero_fiche_exists'
                })

            intervention.numero_fiche = numero_fiche_complet
        except ValueError:
            return JsonResponse({'success': False, 'message': 'Le numéro de fiche doit être un nombre entier'})
        if isModification:
            date = request.POST.get('date')
            date = timezone.make_aware(datetime.strptime(date, '%d/%m/%Y'))
            heure_debut = request.POST.get('heure_debut')
            heure_debut = timezone.make_aware(datetime.strptime(heure_debut, '%H:%M'))
            date_debut = timezone.make_aware(datetime.combine(date.date(), heure_debut.time()))
            intervention.top_debut = date_debut

        # Le reste du code pour mettre à jour l'intervention
        intervention.is_done = True
        intervention.etat_doleance = statut_final
        intervention.description_panne = request.POST.get('description_panne', '').upper()
        intervention.resolution = request.POST.get('resolution', '').upper()
        intervention.observations = request.POST.get('observations', '').upper()
        intervention.pieces_changees = request.POST.get('pieces_changees', '').upper()
        new_element = request.POST.get('element', '').upper()
        intervention.top_terminer = timezone.now()
        heure_fin = request.POST.get('heure_fin')
        if not heure_fin:
            return JsonResponse({'success': False, 'message': 'L\'heure de fin est requise'})

        # Convertir l'heure de fin en objet datetime
        heure_fin = timezone.make_aware(datetime.strptime(heure_fin, '%H:%M'))

        # Utiliser l'heure de fin fournie pour calculer top_terminer
        date_aujourdhui = timezone.now().date()
        if not isModification:
            intervention.top_terminer = timezone.make_aware(datetime.combine(date_aujourdhui, heure_fin.time()))
        else:
            intervention.top_terminer = timezone.make_aware(datetime.combine(date.date(), heure_fin.time()))
        # Calculer la durée de l'intervention
        if intervention.top_debut:
            duree = intervention.top_terminer - intervention.top_debut
            intervention.duree_intervention = int(duree.total_seconds())
        else:
            intervention.duree_intervention = 0

        intervention.save()

        # Mise à jour de la doléance
        doleance = intervention.doleance
        doleance.statut = statut_final
        interventions = Intervention.objects.filter(doleance=doleance)
        doleance.date_debut = interventions.aggregate(Min('top_debut'))['top_debut__min']
        doleance.date_fin = interventions.aggregate(Max('top_terminer'))['top_terminer__max']
        doleance.element = new_element
        doleance.save()
        # Misa à jour Doléance équipes
        DoleanceEquipe.objects.filter(doleance_id=doleance.id).delete()
        # Mise à jour des techniciens
        for intervention_personnel in InterventionPersonnel.objects.filter(intervention=intervention):
            personnel = intervention_personnel.personnel
            personnel.statut = 'PRS'
            personnel.save()

        return JsonResponse({
            'success': True,
            'message': 'Intervention terminée avec succès',
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur inattendue: {str(e)}'})


# #####################Fin Procédure de clôture######################


def get_available_years():
    # Assuming date_transmission is a CharField storing dates in 'YYYY-MM-DD' format
    years_queryset = Doleance.objects.exclude(date_transmission__isnull=True).annotate(
        year=Substr('date_transmission', 1, 4)
    ).values_list('year', flat=True).distinct().order_by('-year')

    # Convert QuerySet to a list of years, ensuring they are integers
    years = [int(year) for year in years_queryset if year.isdigit()]

    # Debug: Print the years to verify
    print("Years queryset:", years)

    # If the queryset is empty, add the current year as a fallback
    if not years:
        current_year = timezone.now().year
        years.append(current_year)

    print("les années:", years)

    return years


# ##################### Début Toules les Doléances ######################

@login_required
def toutes_les_doleances(request):
    current_date = timezone.now()
    current_year = current_date.year
    current_month = current_date.month
    # min_year = Doleance.objects.using('kimei_db').aggregate(Min('date_transmission__year'))[
    #                'date_transmission__year__min'] or current_year
    # max_year = Doleance.objects.using('kimei_db').aggregate(Max('date_transmission__year'))[
    #                'date_transmission__year__max'] or current_year
    # years = range(min_year, max_year + 1)
    years = get_available_years()
    print("les années:", years)

    months = [
        (1, 'Janvier'), (2, 'Février'), (3, 'Mars'), (4, 'Avril'),
        (5, 'Mai'), (6, 'Juin'), (7, 'Juillet'), (8, 'Août'),
        (9, 'Septembre'), (10, 'Octobre'), (11, 'Novembre'), (12, 'Décembre')
    ]
    context = {
        'years': years,
        'months': months,
        'current_year': current_year,
        'current_month': current_month
    }
    return render(request, 'gmao/toutes_les_doleances.html', context)


def get_doleances_data(request):
    current_date = timezone.now()

    year = request.GET.get('year')
    month = request.GET.get('month')
    client_id = request.GET.get('client')

    doleances_query = Doleance.objects.using('kimei_db').exclude(statut='NEW')

    if year and year != 'all':
        doleances_query = doleances_query.filter(date_fin__year=int(year))

    if month and month != 'all':
        if year and year != 'all':

            # Assuming `month` is a variable containing the month number
            # and `current_date` is the current timezone-aware datetime object
            start_date = current_date.replace(month=int(month), day=1, hour=0, minute=0, second=0, microsecond=0)
            end_date = start_date + relativedelta(months=1)
        else:

            # Assuming `month` is a variable containing the month number
            # and `current_date` is the current timezone-aware datetime object
            start_date = current_date.replace(month=int(month), day=1, hour=0, minute=0, second=0, microsecond=0)
            end_date = start_date + relativedelta(months=1)

        doleances_query = doleances_query.filter(date_fin__gte=start_date, date_fin__lt=end_date)

    if client_id:
        doleances_query = doleances_query.filter(station__client_id=client_id)

    doleances = doleances_query.order_by('-date_fin')

    data = []
    for doleance in doleances:
        doleance_data = {
            'id': doleance.id,
            'ndi': doleance.ndi,
            'nombre_interventions': f'<button class="btn btn-primary" onclick="fillIntervention({doleance.id})" data-bs-toggle="modal" data-bs-target="#vue-interventions">'
                                    f'Details <span class="badge bg-dark">'
                                    f'{doleance.intervention_set.all().count()}'
                                    f'</span>'
                                    f'</button>',
            'date_transmission': timezone.localtime(doleance.date_transmission).strftime(
                '%d/%m/%Y %H:%M') if doleance.date_transmission else '',
            'statut': doleance.statut,
            'station': doleance.station.libelle_station if doleance.station else '',
            'element': doleance.element,
            'panne_declarer': doleance.panne_declarer,
            'date_deadline': timezone.localtime(doleance.date_deadline).strftime(
                '%d/%m/%Y %H:%M') if doleance.date_deadline else '',
            'date_debut': timezone.localtime(doleance.date_debut).strftime(
                '%d/%m/%Y %H:%M') if doleance.date_debut else '',
            'date_fin': timezone.localtime(doleance.date_fin).strftime(
                '%d/%m/%Y %H:%M') if doleance.date_fin else '',
            'commentaire': doleance.commentaire,
            # 'actions': f'<button class="btn btn-primary btn-sm view-doleance" data-id="{doleance.id}">Voir</button>'
        }
        data.append(doleance_data)

    return JsonResponse({'data': data})


@login_required
def get_clients(request):
    clients = Client.objects.using('kimei_db').all().values('id', 'nom_client')
    return JsonResponse({'clients': list(clients)})


@login_required
def get_interventions_data(request):
    logger.info("Début de get_interventions_data")
    client_id = request.GET.get('client')
    try:
        today = timezone.localtime(timezone.now()).date()
        start_datetime = timezone.make_aware(datetime.combine(today, time.min))
        end_datetime = timezone.make_aware(datetime.combine(today, time.max))

        if request.user.role == 'ADMIN':
            start_date = request.GET.get('startDate')
            end_date = request.GET.get('endDate')

            if start_date and end_date:
                logger.info(f"Filtrage par date : du {start_date} au {end_date}")
                start_datetime = timezone.make_aware(
                    datetime.combine(datetime.strptime(start_date, '%d/%m/%Y'), time.min))
                end_datetime = timezone.make_aware(datetime.combine(datetime.strptime(end_date, '%d/%m/%Y'), time.max))

        logger.info(f"Filtrage pour la période : du {start_datetime} au {end_datetime}")

        date_filter = (Q(top_depart__range=(start_datetime, end_datetime)) |
                       Q(top_debut__range=(start_datetime, end_datetime)) |
                       Q(top_terminer__range=(start_datetime, end_datetime)))

        if request.user.role == 'ADMIN':
            interventions = Intervention.objects.filter(date_filter)
            if client_id:
                interventions = interventions.filter(doleance__station__client_id=client_id)
        elif request.user.role == 'TECH':
            interventions = Intervention.objects.filter(
                date_filter,
                interventionpersonnel__personnel__matricule=request.user.matricule
            ).distinct()
            if client_id:
                interventions = interventions.filter(doleance__station__client_id=client_id)
        else:
            interventions = Intervention.objects.none()

        interventions = interventions.select_related('doleance', 'doleance__station')
        if not interventions.exists():
            return JsonResponse({'data': []})
        logger.info(f"Nombre d'interventions après filtrage : {interventions.count()}")

        data = []
        for intervention in interventions:
            logger.info(
                f"Dates brutes - ID: {intervention.id}, top_depart: {intervention.top_depart}, top_debut: {intervention.top_debut}, top_terminer: {intervention.top_terminer}")

            # Récupérer les techniciens pour chaque intervention
            techniciens = InterventionPersonnel.objects.filter(intervention=intervention).select_related('personnel')
            techniciens_list = ", ".join(
                [f"{t.personnel.nom_personnel} {t.personnel.prenom_personnel}" for t in techniciens])

            intervention_data = {
                'id': intervention.id,
                'appelant': intervention.doleance.appelant.nom_appelant if intervention.doleance.appelant else '',
                'transmission': intervention.doleance.type_transmission,
                'bt': intervention.doleance.bt if intervention.doleance else '',
                'ndi': intervention.doleance.ndi if intervention.doleance else '',
                'station': intervention.doleance.station.libelle_station if intervention.doleance and intervention.doleance.station else '',
                'element': intervention.doleance.element if intervention.doleance else '',
                'panne': intervention.doleance.panne_declarer if intervention.doleance else '',
                'statut': intervention.etat_doleance if intervention.etat_doleance else '',
                'resolution': intervention.resolution if intervention.resolution else '',
                'date_transmission': intervention.doleance.date_transmission.strftime('%d/%m/%Y %H:%M')
                if intervention.doleance.date_transmission else '',
                'date_deadline': intervention.doleance.date_deadline.strftime('%d/%m/%Y %H:%M')
                if intervention.doleance.date_deadline else '',
                'prise_en_charge': timezone.localtime(intervention.top_depart).strftime(
                    '%d/%m/%Y %H:%M') if intervention.top_depart else '',
                'debut_travail': timezone.localtime(intervention.top_debut).strftime(
                    '%d/%m/%Y %H:%M') if intervention.top_debut else '',
                'fin_travail': timezone.localtime(intervention.top_terminer).strftime(
                    '%d/%m/%Y %H:%M') if intervention.top_terminer else '',
                'numero_fiche': intervention.numero_fiche if intervention.numero_fiche else '',
                'techniciens': techniciens_list,
                'duree_de_travail': str(intervention.duree_intervention) if intervention.duree_intervention else '',
                'kilometrage_depart': str(
                    intervention.kilometrage_depart_debut) if intervention.kilometrage_depart_debut else '',
                'commentaires': intervention.doleance.commentaire if intervention.doleance else '',
                'kilometrage_retour': str(intervention.kilometrage_home) if intervention.kilometrage_home else '',
            }
            data.append(intervention_data)

            # logger.info(f"Intervention ajoutée - ID: {intervention.id}, top_depart: {intervention.top_depart}")
        # logger.info(f"Données complètes avant envoi : {json.dumps(data, default=str)}")
        # logger.info(f"Nombre total d'interventions renvoyées : {len(data)}")
        return JsonResponse({'data': data}, safe=False)

    except Exception as e:
        logger.error(f"Erreur dans get_interventions_data: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Une erreur est survenue lors de la récupération des données'}, status=500)


# ##################### Fin Toutes les Interventions ######################


# #####################Fin Affichage de la liste des interventions######################

@login_required
@require_POST
def prendre_en_charge(request, doleance_id):
    # logger.info(f"Tentative de prise en charge de la doléance {doleance_id}")
    try:
        doleance = get_object_or_404(Doleance.objects.using('kimei_db'), id=doleance_id)
        # logger.info(f"Doléance trouvée: {doleance} {doleance.statut}")

        if doleance.statut not in ['NEW', 'ATP', 'ATD']:
            return JsonResponse({'success': False, 'message': 'Cette doléance ne peut pas être prise en charge'})

        technicien = Personnel.objects.using('kimei_db').get(matricule=request.user.matricule)
        # logger.info(f"Technicien trouvé: {technicien}")

        equipe_personnel = EquipePersonnel.objects.using('teams_db').filter(personnel_id=technicien.id).first()
        # logger.info(f"EquipePersonnel trouvé: {equipe_personnel}")

        if not equipe_personnel:
            return JsonResponse({'success': False, 'message': 'Ce technicien n\'appartient à aucune équipe'})

        equipe = equipe_personnel.equipe
        # logger.info(f"Équipe trouvée: {equipe}")

        # Vérifier si la doléance est déjà associée à une équipe
        # existing_association = DoleanceEquipe.objects.using('teams_db').filter(doleance_id=doleance.id).first()
        # if existing_association:
        #     if existing_association.equipe_id != equipe.id:
        #         # Si la doléance est associée à une autre équipe, la dissocier
        #         existing_association.delete()
        #     else:
        #         # Si la doléance est déjà associée à cette équipe, ne rien faire
        #         logger.info(f"La doléance {doleance_id} est déjà associée à l'équipe {equipe.id}")
        #         return JsonResponse({
        #             'success': True,
        #             'message': 'Cette doléance est déjà prise en charge par votre équipe',
        #             'intervention_id': None
        #         })
        current_time = timezone.now()
        intervention = Intervention.objects.using('kimei_db').create(
            doleance=doleance,
            top_depart=current_time,
            is_done=False,
            is_half_done=False,
            is_going_home=False,
            etat_doleance='ATT'
        )
        # logger.info(f"Intervention créée: {intervention}")

        doleance.statut = 'ATT'
        doleance.save(using='kimei_db')

        membres_equipe = Personnel.objects.using('kimei_db').filter(id__in=equipe.get_personnel_ids())
        for membre in membres_equipe:
            InterventionPersonnel.objects.using('kimei_db').create(intervention=intervention, personnel=membre)
            membre.statut = 'ATT'
            membre.save(using='kimei_db')

        # Créer ou mettre à jour l'association DoleanceEquipe
        DoleanceEquipe.objects.using('teams_db').update_or_create(
            equipe=equipe,
            doleance_id=doleance.id,
            defaults={'equipe': equipe, 'doleance_id': doleance.id}
        )

        return JsonResponse({
            'success': True,
            'message': 'Doléance prise en charge avec succès par l\'équipe',
            'intervention_id': intervention.id,
            'redirect_url': reverse('gmao:detail_intervention', args=[intervention.id])
        })
    except Exception as e:
        logger.error(f"Erreur lors de la prise en charge: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_POST
def annuler_prise_en_charge(request, intervention_id):
    try:

        intervention = get_object_or_404(Intervention, id=intervention_id)
        doleance = intervention.doleance
        ancien_statut = doleance.statut
        if ancien_statut in ['ATP', 'ATD']:
            doleance.statut = ancien_statut
        else:
            doleance.statut = 'NEW'
        doleance.save()

        # Réinitialiser l'intervention
        intervention.delete()

        # Réinitialiser la doléance

        # Réinitialiser le statut des techniciens
        intervention_personnels = InterventionPersonnel.objects.filter(intervention=intervention)
        for intervention_personnel in intervention_personnels:
            technicien = intervention_personnel.personnel
            technicien.statut = 'PRS'
            technicien.save()

        # Supprimer l'association de la doléance à l'équipe
        DoleanceEquipe.objects.filter(doleance=doleance).delete()

        return JsonResponse({
            'success': True,
            'message': 'Prise en charge annulée avec succès',
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@require_POST
@login_required
def affecter_techniciens(request, doleance_id):
    if request.user.role != 'ADMIN':
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)

    data = json.loads(request.body)
    technicien_ids = data.get('techniciens', [])

    try:
        doleance = Doleance.objects.using('kimei_db').get(id=doleance_id)
        techniciens = Personnel.objects.using('kimei_db').filter(id__in=technicien_ids)

        intervention = Intervention.objects.using('kimei_db').create(
            doleance=doleance,
            top_depart=timezone.now(),
            is_done=False,
            is_half_done=False,
            etat_doleance='ATT'
        )

        for technicien in techniciens:
            InterventionPersonnel.objects.using('kimei_db').create(intervention=intervention, personnel=technicien)
            technicien.statut = 'ATT'
            technicien.save()

        doleance.statut = 'ATT'
        doleance.save()

        return JsonResponse({'success': True, 'message': 'Techniciens affectés avec succès'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def get_technicien_portfolio(request):
    if request.user.role != 'TECH':
        return JsonResponse({'success': False, 'message': 'Accès non autorisé'})
    today = timezone.now().date()
    start_of_day = timezone.make_aware(datetime.combine(today, time.min))
    end_of_day = timezone.make_aware(datetime.combine(today, time.max))
    personnel = Personnel.objects.using('kimei_db').get(matricule=request.user.matricule)
    equipe = EquipePersonnel.objects.using('teams_db').filter(personnel_id=personnel.id).first()

    if not equipe:
        return JsonResponse({
            'success': True,
            'doleances': [],
            'message': 'Aucune équipe assignée'
        })

    doleance_ids = list(DoleanceEquipe.objects.using('teams_db')
                        .filter(equipe=equipe.equipe)
                        .values_list('doleance_id', flat=True))

    # Exclure les doléances terminées et les ATP, ATD du jour courant
    doleances = Doleance.objects.using('kimei_db').filter(Q(id__in=doleance_ids) &
                                                          (Q(date_debut__range=(start_of_day, end_of_day)) | ~Q(
                                                              statut='TER'))
                                                          )
    doleances = filter_active_doleances(doleances)

    # Vérifier si une intervention est en cours pour cette équipe
    intervention_en_cours = Intervention.objects.using('kimei_db').filter(
        doleance__id__in=[d.id for d in doleances],
        is_half_done=True,
        is_done=False
    ).exists()

    doleances_data = []
    for d in doleances:
        intervention = Intervention.objects.using('kimei_db').filter(doleance=d).last()
        doleances_data.append({
            'id': d.id,
            'ndi': d.ndi,
            'station': d.station.libelle_station,
            'element': d.element,
            'panne_declarer': d.panne_declarer,
            'statut': d.statut,
            'intervention_id': intervention.id if intervention else None
        })

    return JsonResponse({
        'success': True,
        'doleances': doleances_data,
        'equipe': equipe.equipe.nom if equipe else None,
        'intervention_en_cours': intervention_en_cours
    })


# #####################Début Liste PDR ######################
@login_required
def toutes_les_pieces(request):
    return render(request, 'gmao/toutes_les_pieces.html')


def get_pieces_data(request):
    pieces = Piece.objects.all()

    # Sérialiser les objets Piece en JSON
    pieces_json = serialize('json', pieces)
    pieces_data = json.loads(pieces_json)

    # Transformer les données sérialisées en un format plus simple
    formatted_data = []
    for item in pieces_data:
        piece = item['fields']
        piece['id'] = item['pk']  # Ajouter l'ID à chaque objet

        # Gérer les relations ForeignKey
        if piece['unite']:
            unite = UnitePiece.objects.get(pk=piece['unite'])
            piece['unite'] = unite.unite if hasattr(unite, 'unite') else str(unite)

        if piece['type']:
            type_piece = TypePiece.objects.get(pk=piece['type'])
            piece['type'] = type_piece.libelle_type if hasattr(type_piece, 'libelle_type') else str(type_piece)

        formatted_data.append(piece)

    # Retourner les données au format JSON
    return JsonResponse({'data': formatted_data}, safe=False)


# ##################### Fin Liste PDR ######################


# #####################Début Ajout éléments ######################
#### DEBUT AD ####

@login_required
def liste_appareils_distributeurs(request):
    return render(request, 'gmao/liste_appareils_distributeurs.html')


@csrf_exempt
@login_required
@require_http_methods(["POST"])
def ajouter_distributeur(request):
    try:

        print("Raw request body:", request.body)

        data = json.loads(request.body.decode('utf-8'))
        station = Station.objects.get(id=data['station'])
        modele_ad = ModeleAd.objects.get(id=data['modele_ad'])

        piste = Piste.objects.get(station=station)

        type_contrat = data.get('type_contrat', 'Sous Contrat')

        if type_contrat == 'Sous Contrat':
            type_contrat = 'S'
        elif type_contrat == 'Hors Contrat':
            type_contrat = 'H'
        else:
            type_contrat = 'S'

        print("Parsed JSON data:", data)

        distributeur = AppareilDistribution.objects.create(
            piste=piste,
            modele_ad=modele_ad,
            num_serie=data['num_serie'],
            type_contrat=type_contrat,
            face_principal=data['face_principal'],
            face_secondaire=data['face_secondaire']
        )

        faces_data = data.get('faces', {})
        for face, pistolet_data in faces_data.items():
            Pistolet.objects.create(
                appareil_distribution=distributeur,
                orientation=face[0],  # R ou L
                produit_id=pistolet_data['produit'],
                date_flexible=pistolet_data.get('date_flexible', '')
            )

        return JsonResponse({'success': True, 'message': 'Distributeur ajouté avec succès'})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Données JSON invalides'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@csrf_exempt
@login_required
@require_http_methods(["POST"])
def update_distributeur(request, distributeur_id):
    try:
        data = json.loads(request.body.decode('utf-8'))

        distributeur = AppareilDistribution.objects.get(id=distributeur_id)
        station = Station.objects.get(id=data['station'])

        try:
            piste = Piste.objects.get(station=station)
        except Piste.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'La station sélectionnée n\'a pas de piste associée'})

        type_contrat_frontend = data.get('type_contrat')
        if type_contrat_frontend == 'Sous Contrat':
            type_contrat = 'S'
        elif type_contrat_frontend == 'Hors Contrat':
            type_contrat = 'H'
        else:
            type_contrat = 'S'

        distributeur.piste = piste
        distributeur.modele_ad = ModeleAd.objects.get(id=data['modele_ad'])
        distributeur.num_serie = data['num_serie']
        distributeur.type_contrat = type_contrat

        face_principal = data.get('face_principal')
        face_secondaire = data.get('face_secondaire')

        distributeur.face_principal = int(face_principal) if face_principal else None
        distributeur.face_secondaire = int(face_secondaire) if face_secondaire else None

        if 'annee_installation' in data:
            distributeur.annee_installation = data['annee_installation']
        if 'is_face_unique' in data:
            distributeur.is_face_unique = data['is_face_unique']
        distributeur.save()

        # Supprimer les anciens pistolets
        distributeur.pistolet_set.all().delete()

        # Ajouter les nouveaux pistolets
        faces_data = data.get('faces', {})
        for face, pistolet_data in faces_data.items():
            Pistolet.objects.create(
                appareil_distribution=distributeur,
                orientation=face[0],  # R ou L
                produit_id=pistolet_data['produit'],
                date_flexible=pistolet_data.get('date_flexible', ''),
                type_contrat=type_contrat
            )

        return JsonResponse({'success': True, 'message': 'Distributeur mis à jour avec succès'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_http_methods(["POST"])
def supprimer_distributeur(request, distributeur_id):
    try:
        distributeur = AppareilDistribution.objects.get(id=distributeur_id)
        distributeur.delete()
        return JsonResponse({'success': True, 'message': 'Distributeur supprimé avec succès'})
    except AppareilDistribution.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Distributeur non trouvé'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
def get_distributeur(request, distributeur_id):
    try:
        distributeur = AppareilDistribution.objects.get(id=distributeur_id)
        pistolets = distributeur.pistolet_set.all().order_by('orientation')

        data = {
            'id': distributeur.id,
            'station': distributeur.piste.station.id,
            'modele_ad': distributeur.modele_ad.id,
            'num_serie': distributeur.num_serie,
            'type_contrat': distributeur.type_contrat,
            'face_principal': distributeur.face_principal,
            'face_secondaire': distributeur.face_secondaire,
            'pistolets': [
                {
                    'id': p.id,
                    'orientation': p.orientation,
                    'produit': p.produit.id,
                    'date_flexible': p.date_flexible
                } for p in pistolets
            ]
        }
        return JsonResponse(data)
    except AppareilDistribution.DoesNotExist:
        return JsonResponse({'error': 'Distributeur non trouvé'}, status=404)


@login_required
def get_appareils_distributeurs_data(request):
    user = request.user

    if user.role == 'TECH':
        # Obtenir les doléances du technicien
        personnel = Personnel.objects.using('kimei_db').get(matricule=user.matricule)
        equipe = EquipePersonnel.objects.using('teams_db').filter(personnel_id=personnel.id).first()

        if equipe:
            # Forcer l'évaluation de la requête interne
            doleance_ids = list(DoleanceEquipe.objects.using('teams_db')
                                .filter(equipe=equipe.equipe)
                                .values_list('doleance_id', flat=True))

            doleances = Doleance.objects.using('kimei_db').filter(id__in=doleance_ids)
            doleances = filter_active_doleances(doleances)

            # Forcer l'évaluation de station_ids
            station_ids = list(doleances.values_list('station_id', flat=True).distinct())

            appareils = AppareilDistribution.objects.using('kimei_db').filter(
                piste__station_id__in=station_ids).select_related('piste__station', 'modele_ad')
        else:
            return JsonResponse({"data": []})
    else:
        appareils = AppareilDistribution.objects.using('kimei_db').all().select_related('piste__station', 'modele_ad')

    data = []
    for appareil in appareils:
        pistolets = appareil.pistolet_set.all().order_by('orientation')
        pistolets_info = [
            f"{appareil.num_serie}{p.orientation}-{p.produit.code_produit}-{appareil.type_contrat[0]} *{p.date_flexible or 'ILLI'}*"
            for index, p in enumerate(pistolets)
        ]
        appareil_data = {
            'id': appareil.id,
            'piste': f"{appareil.piste.station.libelle_station} - AD {appareil.face_principal}/{appareil.face_secondaire}",
            'modele_ad': str(appareil.modele_ad),
            'num_serie': appareil.num_serie,
            'type_contrat': appareil.type_contrat,
            'pistolets': '<br>'.join(pistolets_info)
        }
        data.append(appareil_data)
    return JsonResponse({"data": data}, safe=False)


#### FIN AD ####

#### DEBUT CUVE ####

@login_required
def liste_cuves(request):
    return render(request, 'gmao/liste_cuves.html')


#### FIN CUVE ####

@login_required
def get_stations(request):
    stations = Station.objects.all().values('id', 'client__nom_client', 'province_station', 'libelle_station',
                                            'type_contrat')
    return JsonResponse(list(stations), safe=False)


@login_required
def get_modeles_ad(request):
    modeles = ModeleAd.objects.all().values('id', 'libelle_modele')
    return JsonResponse(list(modeles), safe=False)


@login_required
def get_produits(request):
    produits = Produit.objects.all().values('id', 'nom_produit')
    return JsonResponse(list(produits), safe=False)


# #####################Fin Ajout éléments ######################

# Liste Stations
@login_required
def liste_stations(request):
    return render(request, 'gmao/liste_stations.html')


@login_required
def generate_devis(request, doleance_id):
    try:
        doleance = Doleance.objects.get(id=doleance_id)
        filename = doleance.ndi.replace("/", "-")
        filename = filename + "-" + doleance.station.libelle_station + ".xlsx"
        devis = os.path.join(settings.STATIC_URL, 'files/devis.xlsx')
        wb = load_workbook(devis)
        ws = wb.active
        ws['C3'] = doleance.ndi
        ws['C6'] = doleance.station.client.nom_client + " / " + doleance.station.libelle_station
        ws['C7'] = doleance.panne_declarer
        match doleance.type_contrat:
            case "S":
                ws['J3'] = "Sous Contrat"
            case "H":
                ws['J3'] = "Hors Contrat"
            case "D":
                ws['J3'] = "Devis"
            case "C":
                ws['J3'] = "Conso"
            case "P":
                ws['J3'] = "Préventive"
            case _:
                ws['J3'] = "Erreur"
        deadline = doleance.date_deadline.strftime('%d/%m/%Y %H:%M')
        ws['C8'] = deadline
        wb.save(devis)
        if os.path.exists(devis):
            with open(devis, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                response['Content-Disposition'] = 'inline; filename=' + os.path.basename(filename)
                return response
        raise Http404
    except Doleance.DoesNotExist:
        return JsonResponse({'error': 'Doleance erreur'}, status=404)
    # filename = doleance.ndi
    # logger.info(filename)

    # if os.path.exists(file_path):
    #    with open(file_path, 'rb') as fh:
    #        response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
    #        response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
    #        return response
    # raise Http404


@login_required
def interventions_doleance(request, doleance_id):
    try:
        if doleance_id:
            interventions = Intervention.objects.filter(doleance_id=doleance_id)
            ndi = Doleance.objects.get(id=doleance_id).ndi
        else:
            interventions = Intervention.objects.none()

        # interventions = interventions.select_related('doleance', 'doleance__station')
        if not interventions.exists():
            return JsonResponse({'data': []})

        data = []
        for intervention in interventions:
            # Récupérer les techniciens pour chaque intervention
            techniciens_intervention = InterventionPersonnel.objects.filter(intervention=intervention).select_related(
                'personnel')
            techniciens_list = ", ".join(
                [f"{t.personnel.nom_personnel} {t.personnel.prenom_personnel}" for t in techniciens_intervention])

            intervention_data = {
                # 'id': intervention.id,
                # 'appelant': intervention.doleance.appelant.nom_appelant if intervention.doleance.appelant else '',
                # 'bt': intervention.doleance.bt if intervention.doleance else '',
                # 'ndi': intervention.doleance.ndi if intervention.doleance else '',
                # 'station': intervention.doleance.station.libelle_station if intervention.doleance and intervention.doleance.station else '',
                # 'element': intervention.doleance.element if intervention.doleance else '',
                'statut': intervention.etat_doleance if intervention.etat_doleance else '',
                'resolution': intervention.resolution if intervention.resolution else '',
                # 'date_transmission': intervention.doleance.date_transmission.strftime('%d/%m/%Y %H:%M')
                # if intervention.doleance.date_transmission else '',
                # 'date_deadline': intervention.doleance.date_deadline.strftime('%d/%m/%Y %H:%M')
                # if intervention.doleance.date_deadline else '',
                'depart': timezone.localtime(intervention.top_depart).strftime(
                    '%d/%m/%Y %H:%M') if intervention.top_depart else '',
                'debut': timezone.localtime(intervention.top_debut).strftime(
                    '%d/%m/%Y %H:%M') if intervention.top_debut else '',
                'fin': timezone.localtime(intervention.top_terminer).strftime(
                    '%d/%m/%Y %H:%M') if intervention.top_terminer else '',
                'numero_fiche': intervention.numero_fiche if intervention.numero_fiche else '',
                'techniciens': techniciens_list,
                'duree_de_travail': str(intervention.duree_intervention) if intervention.duree_intervention else '',
                'kilometrage_depart': str(
                    intervention.kilometrage_depart_debut) if intervention.kilometrage_depart_debut else '',
                'commentaires': intervention.doleance.commentaire if intervention.doleance else '',
                'kilometrage_retour': str(intervention.kilometrage_home) if intervention.kilometrage_home else '',
            }
            data.append(intervention_data)

        return JsonResponse({'data': data, 'ndi': ndi})

    except Exception as e:
        logger.error(f"Erreur dans get_interventions_data: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Une erreur est survenue lors de la récupération des données'}, status=500)
